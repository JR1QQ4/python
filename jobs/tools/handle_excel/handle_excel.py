#!/usr/bin/python
# -*- coding:utf-8 -*-
import openpyxl, pprint

print('Opening workbook...')

# 1.读取电子表格数据
wb = openpyxl.load_workbook('censuspopdata.xlsx')
sheet = wb['Population by Census Tract']  # print(wb.sheetnames)
county_data = {}

print('Reading rows...')
for row in range(2, sheet.max_row + 1):
    state = sheet['B' + str(row)].value
    county = sheet['C' + str(row)].value
    populations = sheet['D' + str(row)].value

    # 2.填充数据结构
    # {'AK': {'Aleutians East': {'pop': 3141, 'tracts': 1},
    #         'Aleutians West': {'pop': 5561, 'tracts': 2},
    #         'Anchorage': {'pop': 291826, 'tracts': 55},
    #         'Bethel': {'pop': 17013, 'tracts': 3},
    #         'Bristol Bay': {'pop': 997, 'tracts': 1},...
    county_data.setdefault(state, {})
    county_data[state].setdefault(county, {'tracts': 0, 'populations': 0})
    county_data[state][county]['tracts'] += 1
    county_data[state][county]['populations'] += int(populations)

# 3.将结果写入文件
print('Writing results...')
result_file = open('census2010.py', 'w')
result_file.write('all_data = ' + pprint.pformat(county_data))
result_file.close()
print('Done.')
