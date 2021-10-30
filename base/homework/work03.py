#!/usr/bin/python
# -*- coding:utf-8 -*-
import unittest

import openpyxl
from openpyxl.workbook.workbook import Worksheet

# 方法一
# # 1.加载 excel 文件为工作簿对象
# wb = openpyxl.load_workbook('')
# # 获取所有的表单名
# # print(wb.sheetnames)
# # 2.选中表单
# sh: Worksheet = wb['Sheet1']
# # 3.读取数据
# cl = sh.cell(row=1, column=1)
# print(cl.value)

# 方法二
# wb = openpyxl.load_workbook('')
# sh: Worksheet = wb['Sheet1']
# # rows：按行获取表单中所有的格子，每一行的格子放到一个元组中
# rows = list(sh.rows)
# for i in rows:
#     print(i)
# # columns：按列获取表单中所有的格子，每一列的格子放到一个元组中
# columns = list(sh.columns)
# for j in columns:
#     print(j)
# # 获取所有数据
# title = [row.value for row in rows[0]]
# cases = []
# for item in rows[1:]:
#     data = [i.value for i in item]
#     temp = dict(zip(title, data))
#     cases.append(temp)
# print(cases)
from unittestreport import ddt, list_data

from base.homework.work01 import login_check


class HandleExcel:
    def __init__(self, filename, sheetname):
        self.filename = filename
        self.sheetname = sheetname

    def read_date(self):
        wb = openpyxl.load_workbook(self.filename)
        sh = wb[self.sheetname] if self.sheetname is not None else wb[wb.sheetnames[0]]
        res = list(sh.rows)
        title = [i.value for i in res[0]]
        cases = []
        for item in res[1:]:
            data = [i.value for i in item]
            dic = dict(zip(title, data))
            cases.append(dic)
        return cases

    def write_data(self, row, column, val):
        wb = openpyxl.load_workbook(self.filename)
        sh = wb[self.sheetname] if self.sheetname is not None else wb[wb.sheetnames[0]]
        sh.cell(row=row, column=column, value=val)
        wb.save(self.filename)


@ddt
class TestLogin(unittest.TestCase):
    excel = HandleExcel('', '')
    cases = excel.read_date()

    @list_data(cases)
    def test_login(self, item):
        expected = eval(item['expected'])
        params = eval(item['data'])
        row = item['case_id'] + 1
        res = login_check(**params)
        try:
            self.assertEqual(expected, res)
        except AssertionError as e:
            self.excel.write_data(row=row, column=5, val='未通过')
            raise e
        else:
            self.excel.write_data(row=row, column=5, val='通过')


if __name__ == '__main__':
    excel = HandleExcel('', '')
    cases = excel.read_date()
    print(cases)













