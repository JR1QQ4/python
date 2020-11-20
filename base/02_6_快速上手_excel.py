#!/usr/bin/python
# -*- coding:utf-8 -*-
# 第二部分 自动化任务

print('********* 第 12 章 处理 Excel 电子表格 ***********')

import openpyxl

# 使用 openyxl 模块打开 Excel 文档
wb = openpyxl.load_workbook('data\\example.xlsx')
# print(type(wb))  # <class 'openpyxl.workbook.workbook.Workbook'>

# 从工作簿中取得工作表
# print(wb.get_sheet_names())  # 有警告
# print(wb.sheetnames)  # ['Sheet1', 'Sheet2', 'Sheet3']  # 建议使用

# sheet1 = wb.get_sheet_by_name('Sheet1')  # 有警告
# sheet3 = wb['Sheet3']  # 建议使用
# print(sheet3)  # <Worksheet "Sheet3">
# print(type(sheet3))  # <class 'openpyxl.worksheet.worksheet.Worksheet'>
# print(sheet3.title)  # Sheet3

# another_sheet = wb.get_active_sheet()  # 此方法已经没有了
# another_sheet = wb.active
# print(another_sheet)  # <Worksheet "Sheet1">

# 从表中获取单元格
# sheet1 = wb['Sheet1']
# print(sheet1['A1'])  # <Cell 'Sheet1'.A1>
# print(sheet1['A1'].value)  # 4/5/2015 1:34:02 PM
# c = sheet1['B1']
# print(c.value)
# print('Row ' + str(c.row) + ', Column ' + str(c.column) + ' is ' + c.value)
# print('Cell ' + c.coordinate + ' is ' + c.value)
# print(sheet1['C1'].value)
# print(sheet1.cell(row=1, column=2))  # <Cell 'Sheet1'.B1>
# print(sheet1.cell(row=1, column=2).value)  # Apples
# for i in range(1, 8, 2):
#     print(i, sheet1.cell(row=i, column=2).value)

# 获取 Sheet 最大的行数和列数
# sheet = wb['Sheet1']
# # print(sheet.get_highest_row())  # 错误
# print(sheet.max_row)  # 7
# print(sheet.max_column)  # 3

# 遍历打印 sheet 中每行的值
# for row in sheet1.rows:
#     for cell in row:
#         print(cell.value)

# 列字母和数字之间的转换
# from openpyxl.utils import get_column_letter, column_index_from_string
# print(get_column_letter(1))  # A
# print(get_column_letter(2))  # B
# print(get_column_letter(27))  # AA
# print(get_column_letter(900))  # AHP
# sheet = wb['Sheet1']
# print(get_column_letter(sheet.max_column))  # C
# print(column_index_from_string('A'))  # 1
# print(column_index_from_string('AA'))  # 27

# 从表中取得行和列
# sheet = wb['Sheet1']
# print(tuple(sheet['A1':'C3']))  # 获取 A1 到 C3 的单元格
# # ((<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.B1>, <Cell 'Sheet1'.C1>),
# # (<Cell 'Sheet1'.A2>, <Cell 'Sheet1'.B2>, <Cell 'Sheet1'.C2>),
# # (<Cell 'Sheet1'.A3>, <Cell 'Sheet1'.B3>, <Cell 'Sheet1'.C3>))
# for row_of_cell_objects in sheet['A1':'C3']:
#     for cell_obj in row_of_cell_objects:
#         print(cell_obj.coordinate, cell_obj.value)
#     print('--- END OF ROW ---')

# sheet = wb.active
# print(sheet.columns)  # <generator object Worksheet._cells_by_col at 0x0000026F28AB1B48>
# print(list(sheet.columns))
# # [(<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.A2>, <Cell 'Sheet1'.A3>, <Cell 'Sheet1'.A4>,
# # <Cell 'Sheet1'.A5>, <Cell 'Sheet1'.A6>, <Cell 'Sheet1'.A7>),
# # (<Cell 'Sheet1'.B1>, <Cell 'Sheet1'.B2>, <Cell 'Sheet1'.B3>, <Cell 'Sheet1'.B4>,
# # <Cell 'Sheet1'.B5>, <Cell 'Sheet1'.B6>, <Cell 'Sheet1'.B7>),
# # (<Cell 'Sheet1'.C1>, <Cell 'Sheet1'.C2>, <Cell 'Sheet1'.C3>, <Cell 'Sheet1'.C4>,
# # <Cell 'Sheet1'.C5>, <Cell 'Sheet1'.C6>, <Cell 'Sheet1'.C7>)]
# for column_object in sheet.columns:  # 打印所有列
#     for cell in column_object:
#         print(cell.value)
#     print('--- column end ---')











